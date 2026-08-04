import datetime
from fastapi import APIRouter, Request, Depends, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
import os

from app.database import get_db
from app.models import AgendaEvent, Aviso, QuadroAviso, Membro, ConsultaPrivada, ConfiguracaoSistema, TransacaoFinanceira

# Setup router
router = APIRouter()

# Setup templates directory (located in the project root /templates)
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))
def _get_logo_ver():
    try:
        from app.main import get_logo_version
        return get_logo_version()
    except Exception:
        return 1

def _get_logo_uri():
    try:
        from app.main import get_logo_data_uri
        return get_logo_data_uri()
    except Exception:
        return "/static/images/logo.png"

templates.env.globals["get_logo_version"] = _get_logo_ver
templates.env.globals["get_logo_data_uri"] = _get_logo_uri

# Custom template filters for Portuguese date formatting
PT_MONTHS = {
    1: "JAN", 2: "FEV", 3: "MAR", 4: "ABR", 5: "MAI", 6: "JUN",
    7: "JUL", 8: "AGO", 9: "SET", 10: "OUT", 11: "NOV", 12: "DEZ"
}

def format_date(value, format="%d/%m/%Y"):
    if hasattr(value, "strftime"):
        return value.strftime(format)
    return str(value)

def format_day(value):
    if hasattr(value, "day"):
        return f"{value.day:02d}"
    return str(value)

def format_month(value):
    if hasattr(value, "month"):
        return PT_MONTHS.get(value.month, "DIA")
    return "DIA"

templates.env.filters["format_date"] = format_date
templates.env.filters["format_day"] = format_day
templates.env.filters["format_month"] = format_month

def file_exists(filepath: str) -> bool:
    return os.path.exists(filepath)

templates.env.globals["file_exists"] = file_exists

@router.get("/robots.txt", include_in_schema=False)
async def robots():
    return FileResponse("static/robots.txt")


# High-quality sample data to use as a fallback if the database tables are not yet created
DEFAULT_EVENTS = [
    {
        "title": "Gira de Caboclo e Boiadeiro",
        "description": "Linha de Caboclos (trabalhos de cura e passes) e Boiadeiros (limpeza espiritual).",
        "date": datetime.date.today() + datetime.timedelta(days=2),
        "time": "19:30",
        "type": "Gira Pública",
        "status": "Confirmada",
        "publico": True,
        "segmento": "Umbanda"
    },
    {
        "title": "Atendimento Fraterno & Passes",
        "description": "Sessão individual de escuta fraterna e passes de harmonização espiritual.",
        "date": datetime.date.today() + datetime.timedelta(days=5),
        "time": "18:00",
        "type": "Atendimento",
        "status": "Confirmada",
        "publico": True,
        "segmento": "Umbanda"
    },
    {
        "title": "Estudo Doutrinário da Umbanda",
        "description": "Estudo teórico sobre os fundamentos, orixás e a ritualística da Umbanda.",
        "date": datetime.date.today() + datetime.timedelta(days=9),
        "time": "20:00",
        "type": "Estudos",
        "status": "Especial",
        "publico": False,
        "segmento": "Umbanda"
    },
    {
        "title": "Gira de Pretos Velhos",
        "description": "Conselhos benfazejos, passes de desobsessão e a sabedoria dos Pretos Velhos.",
        "date": datetime.date.today() + datetime.timedelta(days=12),
        "time": "19:30",
        "type": "Gira Pública",
        "status": "Confirmada",
        "publico": True,
        "segmento": "Umbanda"
    },
    {
        "title": "Toque de Orixá - Xangô",
        "description": "Celebração e louvação ao Orixá Xangô na tradição de Nação.",
        "date": datetime.date.today() + datetime.timedelta(days=15),
        "time": "19:00",
        "type": "Trabalho Especial",
        "status": "Confirmada",
        "publico": True,
        "segmento": "Candomblé"
    },
    {
        "title": "Ritual Interno de Amaci",
        "description": "Ritual de firmeza de cabeça e amaci exclusivo para médiuns da corrente.",
        "date": datetime.date.today() + datetime.timedelta(days=18),
        "time": "16:00",
        "type": "Sessão Fechada",
        "status": "Confirmada",
        "publico": False,
        "segmento": "Candomblé"
    }
]

DEFAULT_AVISOS = [
    {
        "title": "Retorno das Giras Públicas",
        "content": "O Oloroke Birigui informa que nossas giras e atividades públicas retornaram aos horários normais. Lembre-se de chegar com antecedência para a retirada de senhas.",
        "date_posted": datetime.date.today() - datetime.timedelta(days=1)
    },
    {
        "title": "Estudos para Novos Médiuns",
        "content": "Estão abertas as inscrições para o ciclo de estudos doutrinários de 2026. Se você deseja conhecer melhor a doutrina umbandista ou desenvolver sua mediunidade, procure a secretaria após a próxima gira.",
        "date_posted": datetime.date.today() - datetime.timedelta(days=4)
    },
    {
        "title": "Doações de Agasalhos e Alimentos",
        "content": "Iniciamos nossa Campanha de Inverno. Estamos arrecadando agasalhos, cobertores e alimentos não perecíveis para doação às famílias assistidas por nossa assistência social. Colabore com o que puder!",
        "date_posted": datetime.date.today() - datetime.timedelta(days=8)
    }
]


@router.get("/", response_class=HTMLResponse)
async def home(request: Request, db: Session = Depends(get_db)):
    # Try fetching announcements from database, fallback to sample data if database/table is missing
    try:
        avisos = db.query(Aviso).filter(Aviso.is_active == True).order_by(Aviso.date_posted.desc()).limit(3).all()
        if not avisos:
            avisos = DEFAULT_AVISOS
    except Exception:
        # Graceful fallback when tables are not yet created on Supabase
        avisos = DEFAULT_AVISOS

    try:
        proximos_eventos = db.query(AgendaEvent).filter(
            AgendaEvent.date >= datetime.date.today(),
            AgendaEvent.publico == True
        ).order_by(AgendaEvent.date.asc()).limit(2).all()
        if not proximos_eventos:
            proximos_eventos = [e for e in DEFAULT_EVENTS if e.get("publico", True)][:2]
    except Exception:
        proximos_eventos = [e for e in DEFAULT_EVENTS if e.get("publico", True)][:2]

    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={
            "active_page": "inicio",
            "avisos": avisos,
            "proximos_eventos": proximos_eventos
        }
    )


@router.get("/sobre", response_class=HTMLResponse)
async def sobre(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="sobre.html",
        context={"active_page": "sobre"}
    )


@router.get("/avisos", response_class=HTMLResponse)
async def avisos_public_page(request: Request, db: Session = Depends(get_db)):
    try:
        avisos = db.query(Aviso).filter(Aviso.is_active == True).order_by(Aviso.date_posted.desc()).all()
        if not avisos:
            avisos = DEFAULT_AVISOS
    except Exception:
        avisos = DEFAULT_AVISOS

    return templates.TemplateResponse(
        request=request,
        name="avisos.html",
        context={
            "active_page": "avisos",
            "avisos": avisos
        }
    )


@router.get("/agenda", response_class=HTMLResponse)
async def agenda(request: Request, db: Session = Depends(get_db)):
    import json
    try:
        eventos = db.query(AgendaEvent).filter(AgendaEvent.publico == True).order_by(AgendaEvent.date.asc()).all()
        if not eventos:
            eventos = [e for e in DEFAULT_EVENTS if e.get("publico", True)]
    except Exception:
        eventos = [e for e in DEFAULT_EVENTS if e.get("publico", True)]

    eventos_list = []
    for ev in eventos:
        if isinstance(ev, dict):
            eventos_list.append({
                "id": ev.get("id"),
                "title": ev.get("title"),
                "date": str(ev.get("date")),
                "time": ev.get("time"),
                "type": ev.get("type"),
                "status": ev.get("status"),
                "publico": True,
                "segmento": ev.get("segmento", "Umbanda"),
                "description": ev.get("description", "")
            })
        else:
            eventos_list.append({
                "id": ev.id,
                "title": ev.title,
                "date": str(ev.date),
                "time": ev.time,
                "type": ev.type,
                "status": ev.status,
                "publico": bool(ev.publico),
                "segmento": ev.segmento or "Umbanda",
                "description": ev.description or ""
            })

    return templates.TemplateResponse(
        request=request,
        name="agenda.html",
        context={
            "active_page": "agenda", 
            "eventos": eventos,
            "eventos_json": json.dumps(eventos_list)
        }
    )


@router.get("/contato", response_class=HTMLResponse)
async def contato(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="contato.html",
        context={"active_page": "contato"}
    )


@router.post("/contato", response_class=HTMLResponse)
async def enviar_contato(
    request: Request,
    nome: str = Form(...),
    email: str = Form(...),
    assunto: str = Form(...),
    mensagem: str = Form(...)
):
    # Form submission simulation
    # In the future, this can be saved to a 'mensagens' table or sent via email.
    sucesso_msg = f"Obrigado, {nome}! Sua mensagem sobre '{assunto}' foi enviada com sucesso. Entraremos em contato em breve."
    
    return templates.TemplateResponse(
        request=request,
        name="contato.html",
        context={
            "active_page": "contato",
            "success_message": sucesso_msg
        }
    )


@router.get("/consultas", response_class=HTMLResponse)
async def list_consultas_get(request: Request, db: Session = Depends(get_db)):
    # Active members approved by Pai de Santo / Admin for private consultations
    leaders = db.query(Membro).filter(
        Membro.ativo == True,
        Membro.aprovado_consulta_privada == True
    ).all()
        
    time_slots = ["14:00", "14:30", "15:00", "15:30", "16:00", "16:30", "17:00", "17:30", "18:00", "18:30", "19:00"]
    min_date = datetime.date.today().strftime("%Y-%m-%d")
    
    return templates.TemplateResponse(
        request=request,
        name="consultas.html",
        context={
            "active_page": "consultas",
            "authorized_leaders": leaders,
            "time_slots": time_slots,
            "min_date": min_date,
            "error": None,
            "success": False,
            "form_data": None
        }
    )

@router.post("/consultas", response_class=HTMLResponse)
async def book_consulta_post(
    request: Request,
    membro_id: int = Form(...),
    data: str = Form(...),
    horario: str = Form(...),
    nome_cliente: str = Form(...),
    telefone_cliente: str = Form(...),
    email_cliente: str = Form(None),
    observacoes: str = Form(None),
    tipo: str = Form("Consulta"),
    db: Session = Depends(get_db)
):
    # Repopulate list of leaders for form rendering
    leaders = db.query(Membro).filter(
        Membro.ativo == True,
        Membro.aprovado_consulta_privada == True
    ).all()
        
    time_slots = ["14:00", "14:30", "15:00", "15:30", "16:00", "16:30", "17:00", "17:30", "18:00", "18:30", "19:00"]
    min_date = datetime.date.today().strftime("%Y-%m-%d")
    
    form_data = {
        "membro_id": membro_id,
        "data": data,
        "horario": horario,
        "nome_cliente": nome_cliente,
        "telefone_cliente": telefone_cliente,
        "email_cliente": email_cliente,
        "observacoes": observacoes,
        "tipo": tipo
    }
    
    # Selected leader validation
    selected_leader = db.query(Membro).filter(
        Membro.id == membro_id, 
        Membro.ativo == True,
        Membro.aprovado_consulta_privada == True
    ).first()
    if not selected_leader:
        return templates.TemplateResponse(
            request=request,
            name="consultas.html",
            context={
                "active_page": "consultas",
                "authorized_leaders": leaders,
                "time_slots": time_slots,
                "min_date": min_date,
                "error": "O dirigente ou médium selecionado não foi encontrado ou está inativo.",
                "success": False,
                "form_data": form_data
            }
        )
    
    # Date parsing
    try:
        booking_date = datetime.datetime.strptime(data, "%Y-%m-%d").date()
    except ValueError:
        return templates.TemplateResponse(
            request=request,
            name="consultas.html",
            context={
                "active_page": "consultas",
                "authorized_leaders": leaders,
                "time_slots": time_slots,
                "min_date": min_date,
                "error": "A data selecionada é inválida.",
                "success": False,
                "form_data": form_data
            }
        )
        
    # Past date validation
    if booking_date < datetime.date.today():
        return templates.TemplateResponse(
            request=request,
            name="consultas.html",
            context={
                "active_page": "consultas",
                "authorized_leaders": leaders,
                "time_slots": time_slots,
                "min_date": min_date,
                "error": "Não é possível agendar consultas em datas retroativas.",
                "success": False,
                "form_data": form_data
            }
        )
        
    # Unique booking slot collision check
    existing_booking = db.query(ConsultaPrivada).filter(
        ConsultaPrivada.membro_id == membro_id,
        ConsultaPrivada.data == booking_date,
        ConsultaPrivada.horario == horario,
        ConsultaPrivada.status != "Cancelada"
    ).first()
    
    if existing_booking:
        return templates.TemplateResponse(
            request=request,
            name="consultas.html",
            context={
                "active_page": "consultas",
                "authorized_leaders": leaders,
                "time_slots": time_slots,
                "min_date": min_date,
                "error": f"O horário {horario}h do dia {booking_date.strftime('%d/%m/%Y')} já está reservado com {selected_leader.nome}. Por favor, selecione outro horário ou outra data.",
                "success": False,
                "form_data": form_data
            }
        )
        
    # Create the consultation record
    try:
        new_booking = ConsultaPrivada(
            nome_cliente=nome_cliente.strip(),
            telefone_cliente=telefone_cliente.strip(),
            email_cliente=email_cliente.strip() if email_cliente else None,
            data=booking_date,
            horario=horario,
            membro_id=membro_id,
            tipo=tipo,
            status="Pendente",
            observacoes=observacoes.strip() if observacoes else None
        )
        db.add(new_booking)
        db.commit()
        
        booking_info = {
            "nome_cliente": nome_cliente,
            "leader_name": selected_leader.nome,
            "leader_cargo": selected_leader.cargo,
            "data_formatada": booking_date.strftime("%d/%m/%Y"),
            "horario": horario,
            "tipo": tipo
        }
        
        return templates.TemplateResponse(
            request=request,
            name="consultas.html",
            context={
                "active_page": "consultas",
                "authorized_leaders": leaders,
                "time_slots": time_slots,
                "min_date": min_date,
                "error": None,
                "success": True,
                "booking": booking_info,
                "form_data": None
            }
        )
    except Exception as e:
        db.rollback()
        return templates.TemplateResponse(
            request=request,
            name="consultas.html",
            context={
                "active_page": "consultas",
                "authorized_leaders": leaders,
                "time_slots": time_slots,
                "min_date": min_date,
                "error": f"Erro inesperado ao registrar o agendamento: {e}",
                "success": False,
                "form_data": form_data
            }
        )


# ==========================================================================
# ADMINISTRATIVE BACK-OFFICE OPERATIONS
# ==========================================================================

from app.auth_utils import get_session_user_id, generate_auth_token, verify_password

@router.get("/admin/login", response_class=HTMLResponse)
async def admin_login_get(request: Request):
    # If already logged in, redirect to the dashboard
    if get_session_user_id(request):
        return RedirectResponse(url="/admin", status_code=303)
    return templates.TemplateResponse(request=request, name="login.html", context={"error": None})


@router.post("/admin/login")
async def admin_login_post(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    from app.models import User
    
    user = db.query(User).filter(User.username == username).first()
    if not user or not verify_password(password, user.hashed_password):
        return templates.TemplateResponse(
            request=request, 
            name="login.html", 
            context={"error": "Usuário ou senha incorretos."}
        )
    
    if not user.is_active:
        return templates.TemplateResponse(
            request=request, 
            name="login.html", 
            context={"error": "Esta conta de usuário foi inativada."}
        )
        
    if not user.is_approved:
        return templates.TemplateResponse(
            request=request, 
            name="login.html", 
            context={"error": "Sua conta de membro está aguardando aprovação pela administração/Pai de Santo."}
        )
        
    token = generate_auth_token(user.id)
    request.session["user_id"] = user.id
    request.session["username"] = user.username
    
    response = RedirectResponse(url=f"/admin?auth_token={token}", status_code=303)
    response.set_cookie("oloroke_auth_token", token, max_age=18000, httponly=False)
    return response


@router.get("/admin/logout")
async def admin_logout(request: Request):
    request.session.clear()
    response = RedirectResponse(url="/admin/login", status_code=303)
    response.delete_cookie("oloroke_auth_token")
    return response


@router.get("/admin/cadastro_membro", response_class=HTMLResponse)
async def admin_cadastro_membro_get(request: Request):
    if get_session_user_id(request):
        return RedirectResponse(url="/admin", status_code=303)
    return templates.TemplateResponse(request=request, name="cadastro_membro.html", context={"success": False, "error": None, "form_data": None})


@router.post("/admin/cadastro_membro")
async def admin_cadastro_membro_post(
    request: Request,
    full_name: str = Form(...),
    cpf: str = Form(...),
    birth_date: str = Form(...),
    username: str = Form(...),
    password: str = Form(...),
    whatsapp: str = Form(...),
    email: str = Form(...),
    cep: str = Form(""),
    cidade: str = Form(""),
    uf: str = Form(""),
    endereco: str = Form(""),
    numero: str = Form(""),
    complemento: str = Form(""),
    bairro: str = Form(""),
    emergencia_nome: str = Form(""),
    emergencia_parentesco: str = Form(""),
    emergencia_telefone: str = Form(""),
    emergencia_observacao: str = Form(""),
    db: Session = Depends(get_db)
):
    from app.models import User, Membro
    from app.auth_utils import hash_password
    from app.membro_utils import validate_cpf, format_cpf, generate_next_codigo_membro, clean_digits
    import datetime

    form_data = {
        "full_name": full_name,
        "cpf": cpf,
        "birth_date": birth_date,
        "username": username,
        "whatsapp": whatsapp,
        "email": email,
        "cep": cep,
        "cidade": cidade,
        "uf": uf,
        "endereco": endereco,
        "numero": numero,
        "complemento": complemento,
        "bairro": bairro,
        "emergencia_nome": emergencia_nome,
        "emergencia_parentesco": emergencia_parentesco,
        "emergencia_telefone": emergencia_telefone,
        "emergencia_observacao": emergencia_observacao
    }

    try:
        # Validate CPF
        cpf_formatted = format_cpf(cpf)
        if not validate_cpf(cpf_formatted):
            return templates.TemplateResponse(
                request=request,
                name="cadastro_membro.html",
                context={"success": False, "error": "CPF inválido. Verifique os dígitos digitados.", "form_data": form_data}
            )

        # Check CPF uniqueness
        clean_cpf_val = clean_digits(cpf)
        existing_cpf = db.query(Membro).filter(Membro.cpf != None).all()
        for m in existing_cpf:
            if m.cpf and clean_digits(m.cpf) == clean_cpf_val:
                return templates.TemplateResponse(
                    request=request,
                    name="cadastro_membro.html",
                    context={"success": False, "error": "Este CPF já está cadastrado na corrente.", "form_data": form_data}
                )

        # Check collision on username
        existing_username = db.query(User).filter(User.username == username.strip()).first()
        if existing_username:
            return templates.TemplateResponse(
                request=request,
                name="cadastro_membro.html",
                context={"success": False, "error": "Nome de usuário já está em uso.", "form_data": form_data}
            )

        # Check collision on email
        existing_email = db.query(User).filter(User.email == email.strip()).first()
        if existing_email:
            return templates.TemplateResponse(
                request=request,
                name="cadastro_membro.html",
                context={"success": False, "error": "E-mail já está cadastrado.", "form_data": form_data}
            )

        try:
            birth_date_parsed = datetime.datetime.strptime(birth_date, "%Y-%m-%d").date()
        except ValueError:
            return templates.TemplateResponse(
                request=request,
                name="cadastro_membro.html",
                context={"success": False, "error": "Data de nascimento inválida.", "form_data": form_data}
            )

        # Generate sequential member code (M0001, M0002...)
        next_code = generate_next_codigo_membro(db)

        # Create unapproved User with "membro" role
        new_user = User(
            username=username.strip(),
            email=email.strip(),
            hashed_password=hash_password(password),
            full_name=full_name.strip(),
            phone=whatsapp.strip(),
            birth_date=birth_date_parsed,
            role="membro",
            is_approved=False,
            is_active=True,
            is_admin=True # Ensure dashboard access
        )
        db.add(new_user)

        # Create Membro record staged as Inativo / Pendente
        new_membro = Membro(
            codigo_membro=next_code,
            situacao="Inativo", # Wait for admin approval to activate
            nome=full_name.strip(),
            cpf=cpf_formatted,
            data_nascimento=birth_date_parsed,
            data_cadastro=datetime.datetime.utcnow(),
            whatsapp=whatsapp.strip(),
            telefone=whatsapp.strip(),
            email=email.strip(),
            cep=cep.strip(),
            cidade=cidade.strip(),
            uf=uf.strip().upper(),
            endereco=endereco.strip(),
            numero=numero.strip(),
            complemento=complemento.strip(),
            bairro=bairro.strip(),
            emergencia_nome=emergencia_nome.strip(),
            emergencia_parentesco=emergencia_parentesco.strip(),
            emergencia_telefone=emergencia_telefone.strip(),
            emergencia_observacao=emergencia_observacao.strip(),
            data_ingresso=datetime.date.today(),
            cargo="Médium",
            ativo=False,
            responsavel_cadastro="Cadastro Público",
            responsavel_ultima_alteracao="Membro (Auto-cadastro)"
        )
        db.add(new_membro)
        db.commit()

        return templates.TemplateResponse(
            request=request,
            name="cadastro_membro.html",
            context={"success": True, "error": None, "form_data": None}
        )
    except Exception as e:
        db.rollback()
        return templates.TemplateResponse(
            request=request,
            name="cadastro_membro.html",
            context={"success": False, "error": f"Erro ao solicitar cadastro: {e}", "form_data": form_data}
        )


@router.post("/admin/membros/approve/{target_user_id}")
async def admin_membro_approve(
    target_user_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)

    try:
        from app.models import User, Membro
        current_user = db.query(User).filter(User.id == user_id).first()
        if not current_user or current_user.role not in ["programador", "pai_de_santo"]:
            request.session["msg_error"] = "Acesso negado para aprovar membros."
            return RedirectResponse(url="/admin?tab=membros", status_code=303)

        target_user = db.query(User).filter(User.id == target_user_id).first()
        if not target_user:
            request.session["msg_error"] = "Membro não encontrado."
            return RedirectResponse(url="/admin?tab=membros", status_code=303)

        target_user.is_approved = True
        
        # Automatically insert or update the Membro record so they are officially in the list
        existing_membro = db.query(Membro).filter(Membro.email == target_user.email).first()
        if not existing_membro:
            new_membro = Membro(
                nome=target_user.full_name or target_user.username,
                cargo="Médium",
                telefone=target_user.phone,
                email=target_user.email,
                ativo=True
            )
            db.add(new_membro)
        else:
            existing_membro.ativo = True

        db.commit()
        request.session["msg_success"] = f"Solicitação de {target_user.full_name or target_user.username} aprovada com sucesso!"
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao aprovar membro: {e}"

    return RedirectResponse(url="/admin?tab=membros", status_code=303)


@router.post("/admin/membros/reject/{target_user_id}")
async def admin_membro_reject(
    target_user_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)

    try:
        from app.models import User
        current_user = db.query(User).filter(User.id == user_id).first()
        if not current_user or current_user.role not in ["programador", "pai_de_santo"]:
            request.session["msg_error"] = "Acesso negado para recusar solicitações de membros."
            return RedirectResponse(url="/admin?tab=membros", status_code=303)

        target_user = db.query(User).filter(User.id == target_user_id, User.is_approved == False).first()
        if not target_user:
            request.session["msg_error"] = "Solicitação não encontrada."
            return RedirectResponse(url="/admin?tab=membros", status_code=303)

        name = target_user.full_name or target_user.username
        db.delete(target_user)
        db.commit()
        request.session["msg_success"] = f"Solicitação de cadastro de {name} foi recusada e removida."
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao recusar solicitação: {e}"

    return RedirectResponse(url="/admin?tab=membros", status_code=303)


@router.get("/admin", response_class=HTMLResponse)
async def admin_dashboard(
    request: Request,
    tab: str = "inicio",
    search: str = "",
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    # Read flash messages from session
    msg_success = request.session.pop("msg_success", None)
    msg_error = request.session.pop("msg_error", None)
    
    from app.models import User, Membro, TransacaoFinanceira, Aviso, AgendaEvent
    
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user:
        request.session.clear()
        return RedirectResponse(url="/admin/login", status_code=303)
        
    user_role = current_user.role or "secretario"
    
    # Check if they are a Membro approved for private consultations
    is_allowed_membro = False
    if user_role == "membro":
        m_rec = db.query(Membro).filter(Membro.email == current_user.email).first()
        if m_rec and m_rec.aprovado_consulta_privada:
            is_allowed_membro = True

    # Enforce role boundaries for active tab
    if user_role == "tesoureiro":
        if tab not in ["inicio", "financeiro", "configuracoes", "quadro_avisos", "calendario_giras", "site", "mensalidades", "membros"]:
            tab = "financeiro"
    elif user_role == "secretario":
        if tab not in ["inicio", "site", "configuracoes", "quadro_avisos", "consultas", "membros"]:
            tab = "site"
    elif user_role == "membro":
        allowed_tabs = ["inicio", "site", "configuracoes", "quadro_avisos"]
        if is_allowed_membro:
            allowed_tabs.append("consultas")
        if tab not in allowed_tabs:
            tab = "site"
            
    if tab == "mensalidades" and user_role not in ["programador", "pai_de_santo", "tesoureiro"]:
        tab = "inicio"
        
    if tab == "consultas" and user_role not in ["programador", "pai_de_santo", "secretario"] and not is_allowed_membro:
        tab = "inicio"
            
    # Load system configurations
    config_pct = db.query(ConfiguracaoSistema).filter(ConfiguracaoSistema.chave == "porcentagem_casa").first()
    porcentagem_casa = float(config_pct.valor) if config_pct else 30.0

    # Query corresponding Membro record for the logged in user
    current_membro = db.query(Membro).filter(
        (Membro.email == current_user.email) | 
        (Membro.nome == current_user.full_name) | 
        (Membro.nome == current_user.username)
    ).first()

    context_data = {
        "active_tab": tab,
        "msg_success": msg_success,
        "msg_error": msg_error,
        "search_query": search,
        "current_user": current_user,
        "current_membro": current_membro,
        "is_allowed_membro": is_allowed_membro,
        "porcentagem_casa": porcentagem_casa
    }
    
    # Count pending member requests to display badge
    pending_count = 0
    if current_user.role in ["programador", "pai_de_santo"]:
        try:
            pending_count = db.query(User).filter(User.role == "membro", User.is_approved == False).count()
        except Exception:
            pass
    context_data["pending_members_count"] = pending_count
    
    # Module specific loaders
    if tab == "inicio":
        try:
            membros_count = db.query(Membro).filter(Membro.ativo == True).count()
        except Exception:
            membros_count = 0
        try:
            eventos_count = db.query(AgendaEvent).filter(AgendaEvent.date >= datetime.date.today()).count()
        except Exception:
            eventos_count = 0
        try:
            avisos_count = db.query(Aviso).filter(Aviso.is_active == True).count()
        except Exception:
            avisos_count = 0
            
        context_data["quick_stats"] = {
            "membros_ativos": membros_count,
            "proximos_eventos": eventos_count,
            "avisos_ativos": avisos_count
        }
    elif tab == "financeiro":
        transacoes = db.query(TransacaoFinanceira).order_by(TransacaoFinanceira.data.desc()).all()
        total_receitas = sum(t.valor for t in transacoes if t.tipo == "receita")
        total_despesas = sum(t.valor for t in transacoes if t.tipo == "despesa")
        saldo = total_receitas - total_despesas
        
        context_data["transacoes"] = transacoes
        context_data["fin_summary"] = {
            "receitas": total_receitas,
            "despesas": total_despesas,
            "saldo": saldo
        }
    elif tab == "membros":
        query = db.query(Membro)
        if search:
            query = query.filter(Membro.nome.ilike(f"%{search}%"))
        membros = query.order_by(Membro.nome.asc()).all()
        context_data["membros"] = membros
        
        # Unapproved members awaiting approval from Pai de Santo / Admin
        try:
            pending_members = db.query(User).filter(User.role == "membro", User.is_approved == False).order_by(User.created_at.desc()).all()
        except Exception:
            pending_members = []
        context_data["pending_members"] = pending_members
    elif tab == "site":
        import json
        avisos = db.query(Aviso).order_by(Aviso.date_posted.desc()).all()
        agenda_events = db.query(AgendaEvent).order_by(AgendaEvent.date.desc()).all()
        context_data["avisos"] = avisos
        context_data["agenda_events"] = agenda_events
        events_list = []
        for ev in agenda_events:
            events_list.append({
                "id": ev.id,
                "title": ev.title,
                "date": str(ev.date),
                "time": ev.time,
                "type": ev.type,
                "status": ev.status,
                "publico": bool(ev.publico),
                "segmento": ev.segmento or "Umbanda",
                "description": ev.description or ""
            })
        context_data["agenda_events_json"] = json.dumps(events_list)
    elif tab == "quadro_avisos":
        avisos_internos = db.query(QuadroAviso).order_by(QuadroAviso.date_posted.desc()).all()
        avisos = db.query(Aviso).order_by(Aviso.date_posted.desc()).all()
        context_data["avisos_internos"] = avisos_internos
        context_data["avisos"] = avisos
    elif tab == "calendario_giras":
        import json
        agenda_events = db.query(AgendaEvent).order_by(AgendaEvent.date.desc()).all()
        context_data["agenda_events"] = agenda_events
        events_list = []
        for ev in agenda_events:
            events_list.append({
                "id": ev.id,
                "title": ev.title,
                "date": str(ev.date),
                "time": ev.time,
                "type": ev.type,
                "status": ev.status,
                "publico": bool(ev.publico),
                "segmento": ev.segmento or "Umbanda",
                "description": ev.description or ""
            })
        context_data["agenda_events_json"] = json.dumps(events_list)
    elif tab == "usuarios":
        if user_role in ["programador", "pai_de_santo"]:
            context_data["all_users"] = db.query(User).order_by(User.id.asc()).all()
    elif tab == "mensalidades":
        if user_role in ["programador", "pai_de_santo", "tesoureiro"]:
            from app.models import Membro, MensalidadePagamento
            import datetime
            
            # Fetch active members for selection list
            membros = db.query(Membro).filter(Membro.ativo == True).order_by(Membro.nome.asc()).all()
            
            # Get selected member id from query param
            membro_id_str = request.query_params.get("membro_id")
            selected_membro = None
            if membro_id_str and membro_id_str.isdigit():
                m_id = int(membro_id_str)
                selected_membro = db.query(Membro).filter(Membro.id == m_id, Membro.ativo == True).first()
            
            # Get selected year
            year_str = request.query_params.get("year")
            try:
                year = int(year_str) if (year_str and year_str.isdigit()) else datetime.date.today().year
            except Exception:
                year = datetime.date.today().year
                
            # Fetch payment records for this year (filtered by selected member if applicable)
            if selected_membro:
                payments_query = db.query(MensalidadePagamento).filter(MensalidadePagamento.ano == year, MensalidadePagamento.membro_id == selected_membro.id)
                payments = payments_query.all()
            else:
                payments = []
            
            # Create the payments map (using string keys: "membroId_mesNum")
            payments_map = {}
            for p in payments:
                payments_map[f"{p.membro_id}_{p.mes}"] = {
                    "pago": p.pago,
                    "isento": p.isento,
                    "valor": p.valor,
                    "data_pagamento": p.data_pagamento.strftime("%Y-%m-%d") if p.data_pagamento else None,
                    "observacao": p.observacao or "",
                    "id": p.id
                }
                
            # Months of the year
            months_list = [
                (1, "Janeiro"), (2, "Fevereiro"), (3, "Março"), (4, "Abril"),
                (5, "Maio"), (6, "Junho"), (7, "Julho"), (8, "Agosto"),
                (9, "Setembro"), (10, "Outubro"), (11, "Novembro"), (12, "Dezembro")
            ]
            
            # Years list for the dropdown
            current_yr = datetime.date.today().year
            years_list = [current_yr - 1, current_yr, current_yr + 1]
            
            context_data["membros"] = membros
            context_data["selected_membro"] = selected_membro
            context_data["selected_year"] = year
            context_data["payments_map"] = payments_map
            context_data["months_list"] = months_list
            context_data["years_list"] = years_list
        
    elif tab == "consultas":
        if user_role in ["programador", "pai_de_santo", "secretario"] or is_allowed_membro:
            # Load private consultations with searching
            query = db.query(ConsultaPrivada)
            
            # If they are just a regular approved member, restrict to their own consultations!
            if user_role == "membro":
                membro_rec = db.query(Membro).filter(Membro.email == current_user.email).first()
                if membro_rec:
                    query = query.filter(ConsultaPrivada.membro_id == membro_rec.id)
                else:
                    query = query.filter(ConsultaPrivada.id == -1) # return empty
                    
            if search:
                query = query.filter(
                    ConsultaPrivada.nome_cliente.ilike(f"%{search}%") | 
                    ConsultaPrivada.telefone_cliente.ilike(f"%{search}%") |
                    ConsultaPrivada.observacoes.ilike(f"%{search}%")
                )
            consultas = query.order_by(ConsultaPrivada.data.desc(), ConsultaPrivada.horario.desc()).all()
            
            # Create a dictionary of members for rapid lookup in template
            members_map = {m.id: m for m in db.query(Membro).all()}
            
            context_data["consultas"] = consultas
            context_data["members_map"] = members_map
        
    return templates.TemplateResponse(
        request=request,
        name="admin.html",
        context=context_data
    )


@router.get("/admin/membros/ficha/me", response_class=HTMLResponse)
async def admin_membro_ficha_me(request: Request, db: Session = Depends(get_db)):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
    
    from app.models import User, Membro
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user:
        return RedirectResponse(url="/admin/login", status_code=303)

    if current_user.role not in ['programador', 'pai_de_santo', 'tesoureiro', 'secretario']:
        request.session["msg_error"] = "Apenas administradores (a partir do cargo de Tesoureiro) têm permissão para imprimir fichas."
        return RedirectResponse(url="/admin", status_code=303)

    membro = db.query(Membro).filter(
        (Membro.email == current_user.email) | 
        (Membro.nome == current_user.full_name) | 
        (Membro.nome == current_user.username)
    ).first()

    if not membro:
        membro = Membro(
            id=current_user.id,
            nome=current_user.full_name or current_user.username,
            cargo=current_user.role.title(),
            telefone=current_user.phone or "",
            email=current_user.email or "",
            data_ingresso=current_user.created_at.date() if hasattr(current_user, "created_at") and current_user.created_at else datetime.date.today(),
            ativo=current_user.is_active,
            isento_mensalidade=False,
            aprovado_consulta_privada=False,
            observacoes="Cadastro oficial do usuário no sistema."
        )

    today_str = datetime.date.today().strftime("%d/%m/%Y")
    return templates.TemplateResponse(
        request=request,
        name="ficha_impressao.html",
        context={"membro": membro, "today_date": today_str, "current_user": current_user}
    )


@router.get("/admin/membros/ficha/{membro_id}", response_class=HTMLResponse)
async def admin_membro_ficha(membro_id: int, request: Request, db: Session = Depends(get_db)):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)

    from app.models import User, Membro
    from app.membro_utils import mask_cpf, format_cpf
    import datetime

    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user:
        return RedirectResponse(url="/admin/login", status_code=303)

    membro = db.query(Membro).filter(Membro.id == membro_id).first()
    if not membro:
        request.session["msg_error"] = "Membro não encontrado."
        return RedirectResponse(url="/admin?tab=membros", status_code=303)

    # Check user permissions
    can_edit = current_user.role in ["programador", "pai_de_santo", "secretario", "tesoureiro"]
    can_see_restricted = current_user.role in ["programador", "pai_de_santo", "secretario"]

    # CPF privacy: format if authorized, mask otherwise
    if can_see_restricted:
        masked_cpf = format_cpf(membro.cpf)
    else:
        masked_cpf = mask_cpf(membro.cpf)

    msg_success = request.session.pop("msg_success", None)
    msg_error = request.session.pop("msg_error", None)

    return templates.TemplateResponse(
        request=request,
        name="ficha_membro.html",
        context={
            "membro": membro,
            "current_user": current_user,
            "can_edit": can_edit,
            "can_see_restricted": can_see_restricted,
            "masked_cpf": masked_cpf,
            "msg_success": msg_success,
            "msg_error": msg_error
        }
    )


@router.post("/admin/membros/ficha/{membro_id}/update")
async def admin_membro_ficha_update(
    membro_id: int,
    request: Request,
    situacao: str = Form("Ativo"),
    nome: str = Form(...),
    data_nascimento: str = Form(""),
    cpf: str = Form(""),
    whatsapp: str = Form(""),
    email: str = Form(""),
    cep: str = Form(""),
    cidade: str = Form(""),
    uf: str = Form(""),
    endereco: str = Form(""),
    numero: str = Form(""),
    complemento: str = Form(""),
    bairro: str = Form(""),
    emergencia_nome: str = Form(""),
    emergencia_parentesco: str = Form(""),
    emergencia_telefone: str = Form(""),
    emergencia_observacao: str = Form(""),
    data_ingresso: str = Form(""),
    cargo: str = Form(""),
    orixa_cabeca: str = Form(""),
    orixa_adjunto: str = Form(""),
    entidades_linhas: str = Form(""),
    observacoes_internas: str = Form(""),
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)

    from app.models import User, Membro
    from app.membro_utils import format_cpf, generate_next_codigo_membro
    import datetime

    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user or current_user.role not in ["programador", "pai_de_santo", "secretario", "tesoureiro"]:
        request.session["msg_error"] = "Você não tem permissão para editar fichas de membros."
        return RedirectResponse(url=f"/admin/membros/ficha/{membro_id}", status_code=303)

    membro = db.query(Membro).filter(Membro.id == membro_id).first()
    if not membro:
        request.session["msg_error"] = "Membro não encontrado."
        return RedirectResponse(url="/admin?tab=membros", status_code=303)

    try:
        # Update primary identity fields
        membro.situacao = situacao.strip()
        membro.ativo = (situacao.strip() == "Ativo")
        membro.nome = nome.strip()

        if cpf and not str(cpf).startswith("***"):
            membro.cpf = format_cpf(cpf)

        if data_nascimento:
            try:
                membro.data_nascimento = datetime.datetime.strptime(data_nascimento, "%Y-%m-%d").date()
            except ValueError:
                pass

        membro.whatsapp = whatsapp.strip()
        membro.telefone = whatsapp.strip()
        membro.email = email.strip()
        membro.cep = cep.strip()
        membro.cidade = cidade.strip()
        membro.uf = uf.strip().upper()
        membro.endereco = endereco.strip()
        membro.numero = numero.strip()
        membro.complemento = complemento.strip()
        membro.bairro = bairro.strip()

        membro.emergencia_nome = emergencia_nome.strip()
        membro.emergencia_parentesco = emergencia_parentesco.strip()
        membro.emergencia_telefone = emergencia_telefone.strip()
        membro.emergencia_observacao = emergencia_observacao.strip()

        if data_ingresso:
            try:
                membro.data_ingresso = datetime.datetime.strptime(data_ingresso, "%Y-%m-%d").date()
            except ValueError:
                pass

        if cargo:
            membro.cargo = cargo.strip()
        membro.orixa_cabeca = orixa_cabeca.strip()
        membro.orixa_adjunto = orixa_adjunto.strip()
        membro.entidades_linhas = entidades_linhas.strip()

        # Update restricted section if authorized
        if current_user.role in ["programador", "pai_de_santo", "secretario"]:
            membro.observacoes_internas = observacoes_internas.strip()
            membro.observacoes = observacoes_internas.strip()

        # Ensure codigo_membro exists
        if not membro.codigo_membro:
            membro.codigo_membro = generate_next_codigo_membro(db)

        # Audit metadata
        membro.updated_at = datetime.datetime.utcnow()
        membro.responsavel_ultima_alteracao = current_user.full_name or current_user.username

        db.commit()
        request.session["msg_success"] = "Ficha cadastral atualizada com sucesso!"
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao atualizar ficha: {e}"

    return RedirectResponse(url=f"/admin/membros/ficha/{membro_id}", status_code=303)


# --------------------------------------------------------------------------
# A. FINANCE CONTROLLERS
# --------------------------------------------------------------------------

@router.post("/admin/financeiro/add")
async def financeiro_add(
    request: Request,
    descricao: str = Form(...),
    valor: float = Form(...),
    tipo: str = Form(...),
    categoria: str = Form(...),
    data: str = Form(...),
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    from app.models import User, TransacaoFinanceira
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user or current_user.role not in ["programador", "pai_de_santo", "tesoureiro"]:
        request.session["msg_error"] = "Acesso negado para esta operação financeira."
        return RedirectResponse(url="/admin?tab=financeiro", status_code=303)
        
    try:
        import datetime
        data_parsed = datetime.datetime.strptime(data, "%Y-%m-%d").date()
        
        new_tx = TransacaoFinanceira(
            descricao=descricao,
            valor=valor,
            tipo=tipo,
            categoria=categoria,
            data=data_parsed
        )
        db.add(new_tx)
        db.commit()
        request.session["msg_success"] = "Lançamento financeiro registrado com sucesso!"
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao registrar lançamento: {e}"
        
    return RedirectResponse(url="/admin?tab=financeiro", status_code=303)


@router.post("/admin/financeiro/delete/{tx_id}")
async def financeiro_delete(
    tx_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    from app.models import User, TransacaoFinanceira
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user or current_user.role not in ["programador", "pai_de_santo", "tesoureiro"]:
        request.session["msg_error"] = "Acesso negado para esta operação financeira."
        return RedirectResponse(url="/admin?tab=financeiro", status_code=303)
        
    try:
        tx = db.query(TransacaoFinanceira).filter(TransacaoFinanceira.id == tx_id).first()
        if tx:
            db.delete(tx)
            db.commit()
            request.session["msg_success"] = "Lançamento financeiro excluído com sucesso."
        else:
            request.session["msg_error"] = "Lançamento não encontrado."
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao excluir lançamento: {e}"
        
    return RedirectResponse(url="/admin?tab=financeiro", status_code=303)


@router.post("/admin/mensalidades/save")
async def mensalidades_save(
    request: Request,
    membro_id: int = Form(...),
    ano: int = Form(...),
    mes: int = Form(...),
    pago: str = Form("false"),
    isento: str = Form("false"),
    valor: float = Form(...),
    data_pagamento: str = Form(None),
    observacao: str = Form(None),
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    from app.models import User, Membro, MensalidadePagamento, TransacaoFinanceira
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user or current_user.role not in ["programador", "pai_de_santo", "tesoureiro"]:
        request.session["msg_error"] = "Acesso negado. Apenas Programador, Pai de Santo e Tesoureiro(a) podem gerenciar mensalidades."
        return RedirectResponse(url=f"/admin?tab=mensalidades&year={ano}", status_code=303)

    try:
        import datetime
        is_paid = (pago.lower() == "true")
        is_exempt = (isento.lower() == "true")
        
        # Parse payment date
        p_date = None
        if is_paid and data_pagamento:
            try:
                p_date = datetime.datetime.strptime(data_pagamento, "%Y-%m-%d").date()
            except Exception:
                p_date = datetime.date.today()
        elif is_paid:
            p_date = datetime.date.today()

        # Find if payment record already exists
        record = db.query(MensalidadePagamento).filter(
            MensalidadePagamento.membro_id == membro_id,
            MensalidadePagamento.ano == ano,
            MensalidadePagamento.mes == mes
        ).first()

        membro = db.query(Membro).filter(Membro.id == membro_id).first()
        membro_nome = membro.nome if membro else "Membro"

        if record:
            record.pago = is_paid
            record.isento = is_exempt
            record.valor = valor
            record.data_pagamento = p_date if is_paid else None
            record.observacao = observacao.strip() if observacao else None
        else:
            record = MensalidadePagamento(
                membro_id=membro_id,
                ano=ano,
                mes=mes,
                pago=is_paid,
                isento=is_exempt,
                valor=valor,
                data_pagamento=p_date if is_paid else None,
                observacao=observacao.strip() if observacao else None
            )
            db.add(record)

        # Sync with general finance ledger
        tx_desc = f"Mensalidade: {membro_nome} - Mês {mes:02d}/{ano}"
        existing_tx = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.descricao == tx_desc,
            TransacaoFinanceira.categoria == "Mensalidade"
        ).first()

        if is_paid and not is_exempt:
            if existing_tx:
                existing_tx.valor = valor
                existing_tx.data = p_date
            else:
                new_tx = TransacaoFinanceira(
                    descricao=tx_desc,
                    valor=valor,
                    tipo="receita",
                    categoria="Mensalidade",
                    data=p_date
                )
                db.add(new_tx)
        else:
            if existing_tx:
                db.delete(existing_tx)

        db.commit()
        request.session["msg_success"] = f"Mensalidade de {membro_nome} ({mes:02d}/{ano}) atualizada com sucesso!"
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao atualizar mensalidade: {e}"

    return RedirectResponse(url=f"/admin?tab=mensalidades&year={ano}&membro_id={membro_id}", status_code=303)


@router.post("/admin/membros/update-mensalidade/{membro_id}")
async def update_membro_mensalidade(
    membro_id: int,
    request: Request,
    valor_mensalidade: float = Form(...),
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    from app.models import User, Membro
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user or current_user.role not in ["programador", "pai_de_santo", "tesoureiro"]:
        request.session["msg_error"] = "Acesso negado. Apenas Programador, Pai de Santo e Tesoureiro(a) podem gerenciar mensalidades."
        return RedirectResponse(url="/admin?tab=mensalidades", status_code=303)
        
    try:
        membro = db.query(Membro).filter(Membro.id == membro_id).first()
        if membro:
            membro.valor_mensalidade = valor_mensalidade
            db.commit()
            request.session["msg_success"] = f"Valor padrão da mensalidade de {membro.nome} updated to R$ {valor_mensalidade:.2f}."
        else:
            request.session["msg_error"] = "Membro não encontrado."
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao atualizar valor padrão: {e}"
        
    import datetime
    year = request.query_params.get("year", datetime.date.today().year)
    return RedirectResponse(url=f"/admin?tab=mensalidades&year={year}&membro_id={membro_id}", status_code=303)


# --------------------------------------------------------------------------
# B. MEMBER CONTROLLERS
# --------------------------------------------------------------------------

@router.post("/admin/membros/add")
async def membros_add(
    request: Request,
    nome: str = Form(...),
    cargo: str = Form(...),
    telefone: str = Form(None),
    email: str = Form(None),
    data_ingresso: str = Form(...),
    ativo: str = Form("true"),
    observacoes: str = Form(None),
    valor_mensalidade: float = Form(50.0),
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    from app.models import User, Membro
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user or current_user.role not in ["programador", "pai_de_santo"]:
        request.session["msg_error"] = "Acesso negado para gerenciar a corrente."
        return RedirectResponse(url="/admin?tab=membros", status_code=303)
        
    try:
        import datetime
        data_parsed = datetime.datetime.strptime(data_ingresso, "%Y-%m-%d").date()
        is_ativo = ativo == "true"
        
        new_membro = Membro(
            nome=nome,
            cargo=cargo,
            telefone=telefone,
            email=email,
            data_ingresso=data_parsed,
            ativo=is_ativo,
            observacoes=observacoes,
            valor_mensalidade=valor_mensalidade
        )
        db.add(new_membro)
        db.commit()
        request.session["msg_success"] = f"Membro '{nome}' cadastrado com sucesso!"
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao cadastrar membro: {e}"
        
    return RedirectResponse(url="/admin?tab=membros", status_code=303)


@router.post("/admin/membros/toggle/{membro_id}")
async def membros_toggle(
    membro_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    from app.models import User, Membro
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user or current_user.role not in ["programador", "pai_de_santo"]:
        request.session["msg_error"] = "Acesso negado para gerenciar a corrente."
        return RedirectResponse(url="/admin?tab=membros", status_code=303)
        
    try:
        membro = db.query(Membro).filter(Membro.id == membro_id).first()
        if membro:
            membro.ativo = not membro.ativo
            db.commit()
            status_str = "ativado" if membro.ativo else "afastado da corrente"
            request.session["msg_success"] = f"Status do membro '{membro.nome}' alterado para: {status_str}."
        else:
            request.session["msg_error"] = "Membro não encontrado."
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao alterar status: {e}"
        
    return RedirectResponse(url="/admin?tab=membros", status_code=303)


@router.post("/admin/membros/consulta-privada/{membro_id}")
async def membros_consulta_privada(
    membro_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    from app.models import User, Membro
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user or current_user.role not in ["programador", "pai_de_santo"]:
        request.session["msg_error"] = "Você não tem permissão para gerenciar as permissões de consulta privada."
        return RedirectResponse(url="/admin?tab=membros", status_code=303)
        
    try:
        form = await request.form()
        aprovado = form.get("aprovado") == "true"
        valor_str = form.get("valor")
        valor = float(valor_str) if valor_str else 0.0
        
        membro = db.query(Membro).filter(Membro.id == membro_id).first()
        if membro:
            membro.aprovado_consulta_privada = aprovado
            membro.valor_consulta = valor
            db.commit()
            request.session["msg_success"] = f"Configurações de consulta privada do membro '{membro.nome}' atualizadas com sucesso!"
        else:
            request.session["msg_error"] = "Membro não encontrado."
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao atualizar configurações: {e}"
        
    return RedirectResponse(url="/admin?tab=membros", status_code=303)


@router.post("/admin/membros/isentar/{membro_id}")
async def membros_isentar(
    membro_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    from app.models import User, Membro
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user or current_user.role not in ["programador", "pai_de_santo", "tesoureiro"]:
        request.session["msg_error"] = "Você não tem permissão para gerenciar a isenção de mensalidade."
        return RedirectResponse(url="/admin?tab=membros", status_code=303)
        
    try:
        form = await request.form()
        isento = form.get("isento") == "true"
        
        membro = db.query(Membro).filter(Membro.id == membro_id).first()
        if membro:
            membro.isento_mensalidade = isento
            db.commit()
            status_str = "isento(a) de mensalidade" if isento else "sujeito(a) a mensalidade"
            request.session["msg_success"] = f"Membro '{membro.nome}' agora está {status_str}!"
        else:
            request.session["msg_error"] = "Membro não encontrado."
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao atualizar isenção: {e}"
        
    # Redirect back to where the request came from if referer exists, otherwise default to membros tab
    referrer = request.headers.get("referer", "/admin?tab=membros")
    return RedirectResponse(url=referrer, status_code=303)


@router.post("/admin/membros/delete/{membro_id}")
async def membros_delete(
    membro_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    from app.models import User, Membro
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user or current_user.role not in ["programador", "pai_de_santo"]:
        request.session["msg_error"] = "Acesso negado para gerenciar a corrente."
        return RedirectResponse(url="/admin?tab=membros", status_code=303)
        
    try:
        membro = db.query(Membro).filter(Membro.id == membro_id).first()
        if membro:
            nome = membro.nome
            db.delete(membro)
            db.commit()
            request.session["msg_success"] = f"Cadastro do membro '{nome}' excluído com sucesso."
        else:
            request.session["msg_error"] = "Membro não encontrado."
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao excluir membro: {e}"
        
    return RedirectResponse(url="/admin?tab=membros", status_code=303)


@router.get("/admin/membros/export-excel")
async def export_membros_excel(
    request: Request,
    membro_id: int = None,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    from app.models import User, Membro
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user or current_user.role not in ["programador", "pai_de_santo", "secretario", "tesoureiro", "membro"]:
        request.session["msg_error"] = "Acesso negado para exportar relatórios de membros."
        return RedirectResponse(url="/admin?tab=membros", status_code=303)

    # If role is 'membro', limit export to their own profile
    if current_user.role == "membro":
        m_rec = db.query(Membro).filter(Membro.email == current_user.email).first()
        if m_rec:
            membro_id = m_rec.id
        else:
            request.session["msg_error"] = "Sua conta de membro não possui cadastro de ficha associado."
            return RedirectResponse(url="/admin?tab=membros", status_code=303)

    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import Response

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ficha de Membros"

    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0B1F3A", end_color="0B1F3A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Title banner
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "FICHA DE CADASTRO DE MEMBROS - TERREIRO OLOROKE BIRIGUI"
    title_cell.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1E4D8C", end_color="1E4D8C", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Subtitle / info
    ws.merge_cells("A2:K2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Relatório extraído em: {datetime.datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    sub_cell.font = Font(name="Calibri", size=9, italic=True, color="555555")
    sub_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 18

    headers = [
        "Código",
        "Situação",
        "Nome Completo",
        "CPF",
        "Data Nascimento",
        "WhatsApp / Telefone",
        "E-mail",
        "CEP",
        "Cidade / UF",
        "Endereço Completo",
        "Contato Emergência - Nome",
        "Contato Emergência - Fone",
        "Data de Ingresso",
        "Cargo / Função",
        "Orixá de Cabeça",
        "Orixá Adjunto",
        "Entidades / Linhas",
        "Consulta Privada",
        "Mensalidade",
        "Observações Internas",
        "Última Atualização",
        "Resp. Alteração"
    ]

    ws.append([]) # row 3 blank
    ws.append(headers) # row 4
    ws.row_dimensions[4].height = 26

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    query = db.query(Membro)
    if membro_id:
        membros = query.filter(Membro.id == membro_id).all()
    else:
        membros = query.order_by(Membro.codigo_membro.asc(), Membro.nome.asc()).all()

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    for row_idx, m in enumerate(membros, start=5):
        ws.row_dimensions[row_idx].height = 22
        
        data_nasc = m.data_nascimento.strftime("%d/%m/%Y") if m.data_nascimento else "-"
        data_ing = m.data_ingresso.strftime("%d/%m/%Y") if m.data_ingresso else "-"
        cidade_uf = f"{m.cidade}/{m.uf}" if m.cidade and m.uf else (m.cidade or m.uf or "-")
        endereco_comp = f"{m.endereco or ''}, {m.numero or ''} - {m.bairro or ''} ({m.complemento or ''})".strip(" ,-()")
        if not endereco_comp:
            endereco_comp = "-"
        
        emergencia_info = f"{m.emergencia_nome or ''} ({m.emergencia_parentesco or 'Contato'})".strip(" ()")
        if not emergencia_info or emergencia_info == "()":
            emergencia_info = "-"

        consulta_str = "Habilitado" if m.aprovado_consulta_privada else "Não Habilitado"
        mensalidade_str = "Isento" if m.isento_mensalidade else f"R$ {m.valor_mensalidade or 50.0:.2f}"
        updated_str = m.updated_at.strftime("%d/%m/%Y %H:%M") if m.updated_at else "-"

        row_data = [
            m.codigo_membro or f"M{m.id:04d}",
            m.situacao or ("Ativo" if m.ativo else "Inativo"),
            m.nome,
            m.cpf or "-",
            data_nasc,
            m.whatsapp or m.telefone or "-",
            m.email or "-",
            m.cep or "-",
            cidade_uf,
            endereco_comp,
            emergencia_info,
            m.emergencia_telefone or "-",
            data_ing,
            m.cargo or "Médium",
            m.orixa_cabeca or "-",
            m.orixa_adjunto or "-",
            m.entidades_linhas or "-",
            consulta_str,
            mensalidade_str,
            m.observacoes_internas or m.observacoes or "-",
            updated_str,
            m.responsavel_ultima_alteracao or "-"
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in [1, 2, 4, 5, 8, 9, 13, 18, 19, 21]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Column width auto-fit
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"ficha_membros_{filename_part}_{datetime.date.today().strftime('%Y%m%d')}.xlsx"

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"{filename}\""
        }
    )


# --------------------------------------------------------------------------
# C. SITE CONTENT CONTROLLERS (AVISOS & AGENDA)
# --------------------------------------------------------------------------

@router.post("/admin/avisos/add")
async def avisos_add(
    request: Request,
    title: str = Form(...),
    content: str = Form(...),
    pub_site: str = Form(None),
    pub_internal: str = Form(None),
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    try:
        from app.models import User, Aviso, QuadroAviso
        current_user = db.query(User).filter(User.id == user_id).first()
        if not current_user or current_user.role not in ["programador", "pai_de_santo", "secretario"]:
            request.session["msg_error"] = "Acesso negado para cadastrar avisos."
            return RedirectResponse(url="/admin?tab=quadro_avisos", status_code=303)
            
        import datetime
        is_site = pub_site == "true"
        is_internal = pub_internal == "true"
        
        # Se nenhum for selecionado, assume site por padrão
        if not is_site and not is_internal:
            is_site = True
            
        saved_targets = []
        if is_site:
            new_aviso = Aviso(
                title=title,
                content=content,
                date_posted=datetime.date.today(),
                is_active=True
            )
            db.add(new_aviso)
            saved_targets.append("Site Público")
            
        if is_internal:
            new_internal = QuadroAviso(
                title=title,
                content=content,
                author_name=current_user.username,
                date_posted=datetime.date.today()
            )
            db.add(new_internal)
            saved_targets.append("Quadro de Avisos Interno")
            
        db.commit()
        targets_str = " e ".join(saved_targets)
        request.session["msg_success"] = f"Aviso publicado com sucesso no {targets_str}!"
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao publicar aviso: {e}"
        
    return RedirectResponse(url="/admin?tab=quadro_avisos", status_code=303)


@router.post("/admin/avisos/edit/{aviso_id}")
async def avisos_toggle_status(
    aviso_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    try:
        from app.models import User, Aviso
        current_user = db.query(User).filter(User.id == user_id).first()
        if not current_user or current_user.role not in ["programador", "pai_de_santo", "secretario"]:
            request.session["msg_error"] = "Acesso negado para alterar status de avisos."
            return RedirectResponse(url="/admin?tab=quadro_avisos", status_code=303)
            
        aviso = db.query(Aviso).filter(Aviso.id == aviso_id).first()
        if aviso:
            aviso.is_active = not aviso.is_active
            db.commit()
            status_str = "visível no site" if aviso.is_active else "ocultado do site"
            request.session["msg_success"] = f"Aviso '{aviso.title}' alterado para {status_str}."
        else:
            request.session["msg_error"] = "Aviso não encontrado."
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao atualizar aviso: {e}"
        
    return RedirectResponse(url="/admin?tab=quadro_avisos", status_code=303)


@router.post("/admin/avisos/delete/{aviso_id}")
async def avisos_delete(
    aviso_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    try:
        from app.models import User, Aviso
        current_user = db.query(User).filter(User.id == user_id).first()
        if not current_user or current_user.role not in ["programador", "pai_de_santo", "secretario"]:
            request.session["msg_error"] = "Acesso negado para excluir avisos."
            return RedirectResponse(url="/admin?tab=quadro_avisos", status_code=303)
            
        aviso = db.query(Aviso).filter(Aviso.id == aviso_id).first()
        if aviso:
            title = aviso.title
            db.delete(aviso)
            db.commit()
            request.session["msg_success"] = f"Aviso '{title}' removido com sucesso."
        else:
            request.session["msg_error"] = "Aviso não encontrado."
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao excluir aviso: {e}"
        
    return RedirectResponse(url="/admin?tab=quadro_avisos", status_code=303)


@router.post("/admin/agenda/add")
async def agenda_add(
    request: Request,
    title: str = Form(...),
    date: str = Form(...),
    time: str = Form(...),
    type: str = Form(...),
    status: str = Form(...),
    description: str = Form(None),
    publico: str = Form("true"),
    segmento: str = Form("Umbanda"),
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    try:
        from app.models import User, AgendaEvent
        current_user = db.query(User).filter(User.id == user_id).first()
        if not current_user or current_user.role not in ["programador", "pai_de_santo", "secretario", "tesoureiro"]:
            request.session["msg_error"] = "Acesso negado para agendar trabalhos. Apenas administradores e dirigentes têm permissão."
            return RedirectResponse(url="/admin?tab=site", status_code=303)
            
        import datetime
        date_parsed = datetime.datetime.strptime(date, "%Y-%m-%d").date()
        is_public = True if str(publico).lower() in ["true", "on", "1"] else False
        
        new_event = AgendaEvent(
            title=title,
            description=description,
            date=date_parsed,
            time=time,
            type=type,
            status=status,
            publico=is_public,
            segmento=segmento
        )
        db.add(new_event)
        db.commit()
        request.session["msg_success"] = f"Trabalho '{title}' agendado com sucesso!"
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao agendar trabalho: {e}"
        
    return RedirectResponse(url="/admin?tab=site", status_code=303)


@router.post("/admin/agenda/edit/{event_id}")
async def agenda_toggle_status(
    event_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    try:
        from app.models import User, AgendaEvent
        current_user = db.query(User).filter(User.id == user_id).first()
        if not current_user or current_user.role not in ["programador", "pai_de_santo", "secretario", "tesoureiro"]:
            request.session["msg_error"] = "Acesso negado para alterar status de trabalhos."
            return RedirectResponse(url="/admin?tab=site", status_code=303)
            
        event = db.query(AgendaEvent).filter(AgendaEvent.id == event_id).first()
        if event:
            # Cycle status: Confirmada -> Especial -> Suspensa -> Confirmada
            if event.status == "Confirmada":
                event.status = "Especial"
            elif event.status == "Especial":
                event.status = "Suspensa"
            else:
                event.status = "Confirmada"
            db.commit()
            request.session["msg_success"] = f"Status da gira '{event.title}' alterado para '{event.status}'."
        else:
            request.session["msg_error"] = "Gira não encontrada."
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao alterar status da gira: {e}"
        
    return RedirectResponse(url="/admin?tab=site", status_code=303)


@router.post("/admin/agenda/update/{event_id}")
async def agenda_update(
    event_id: int,
    request: Request,
    title: str = Form(...),
    date: str = Form(...),
    time: str = Form(...),
    type: str = Form(...),
    status: str = Form(...),
    description: str = Form(None),
    publico: str = Form("true"),
    segmento: str = Form("Umbanda"),
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    try:
        from app.models import User, AgendaEvent
        current_user = db.query(User).filter(User.id == user_id).first()
        if not current_user or current_user.role not in ["programador", "pai_de_santo", "secretario", "tesoureiro"]:
            request.session["msg_error"] = "Acesso negado para editar trabalhos da agenda."
            return RedirectResponse(url="/admin?tab=site", status_code=303)
            
        import datetime
        event = db.query(AgendaEvent).filter(AgendaEvent.id == event_id).first()
        if event:
            date_parsed = datetime.datetime.strptime(date, "%Y-%m-%d").date()
            is_public = True if str(publico).lower() in ["true", "on", "1"] else False
            event.title = title
            event.date = date_parsed
            event.time = time
            event.type = type
            event.status = status
            event.description = description
            event.publico = is_public
            event.segmento = segmento
            db.commit()
            request.session["msg_success"] = f"Gira '{title}' atualizada com sucesso!"
        else:
            request.session["msg_error"] = "Gira não encontrada."
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao atualizar gira: {e}"
        
    return RedirectResponse(url="/admin?tab=site", status_code=303)


@router.post("/admin/agenda/delete/{event_id}")
async def agenda_delete(
    event_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    try:
        from app.models import User, AgendaEvent
        current_user = db.query(User).filter(User.id == user_id).first()
        if not current_user or current_user.role not in ["programador", "pai_de_santo", "secretario", "tesoureiro"]:
            request.session["msg_error"] = "Acesso negado para excluir trabalhos da agenda."
            return RedirectResponse(url="/admin?tab=site", status_code=303)
            
        event = db.query(AgendaEvent).filter(AgendaEvent.id == event_id).first()
        if event:
            title = event.title
            db.delete(event)
            db.commit()
            request.session["msg_success"] = f"Gira '{title}' removida com sucesso."
        else:
            request.session["msg_error"] = "Gira não encontrada."
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao excluir gira: {e}"
        
    return RedirectResponse(url="/admin?tab=site", status_code=303)


@router.post("/admin/usuario/update")
async def admin_user_update(
    request: Request,
    username: str = Form(...),
    email: str = Form(...),
    password: str = Form(None),
    confirm_password: str = Form(None),
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    try:
        from app.models import User
        from app.auth_utils import hash_password
        
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            request.session["msg_error"] = "Usuário não encontrado."
            return RedirectResponse(url="/admin?tab=configuracoes", status_code=303)
            
        # Check username collision
        existing_user = db.query(User).filter(User.username == username, User.id != user_id).first()
        if existing_user:
            request.session["msg_error"] = "Este nome de usuário já está sendo utilizado por outro administrador."
            return RedirectResponse(url="/admin?tab=configuracoes", status_code=303)
            
        user.username = username.strip()
        user.email = email.strip()
        
        if password:
            if password != confirm_password:
                request.session["msg_error"] = "As senhas digitadas não coincidem."
                return RedirectResponse(url="/admin?tab=configuracoes", status_code=303)
            user.hashed_password = hash_password(password)
            
        db.commit()
        
        # Update session details
        request.session["username"] = user.username
        request.session["msg_success"] = "Perfil do administrador atualizado com sucesso!"
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao atualizar perfil: {e}"
        
    return RedirectResponse(url="/admin?tab=configuracoes", status_code=303)


@router.post("/admin/config/logo")
async def admin_config_logo(
    request: Request,
    logo: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    try:
        from app.models import User
        
        user = db.query(User).filter(User.id == user_id).first()
        if not user or (user.role == "membro" and not user.is_admin):
            request.session["msg_error"] = "Você não tem permissão para alterar as configurações do terreiro."
            return RedirectResponse(url="/admin?tab=configuracoes", status_code=303)
            
        data = None
        filename = "logo.png"
        content_type = "image/png"

        if logo and hasattr(logo, "filename") and logo.filename:
            filename = logo.filename
            content_type = getattr(logo, "content_type", None) or "image/png"
            data = await logo.read()

        if not data:
            form = await request.form()
            f = form.get("logo")
            if f and hasattr(f, "filename") and f.filename:
                filename = f.filename
                content_type = getattr(f, "content_type", None) or "image/png"
                data = await f.read()

        if not data:
            request.session["msg_error"] = "Nenhum arquivo enviado ou arquivo inválido. Por favor selecione uma imagem."
            return RedirectResponse(url="/admin?tab=configuracoes", status_code=303)

        from app.main import set_global_logo_cache, detect_image_mime
        mime_type = detect_image_mime(data, content_type)

        import base64
        b64_str = base64.b64encode(data).decode("utf-8")

        # 1. Save to in-memory cache and local file backups
        try:
            set_global_logo_cache(data, mime_type)
        except Exception as ex:
            print(f"Warning updating logo cache: {ex}")

        # 2. Save to SQLite database (ConfiguracaoSistema) for persistent storage
        try:
            c_b64 = db.query(ConfiguracaoSistema).filter(ConfiguracaoSistema.chave == "logo_b64").first()
            if not c_b64:
                db.add(ConfiguracaoSistema(chave="logo_b64", valor=b64_str))
            else:
                c_b64.valor = b64_str

            c_mime = db.query(ConfiguracaoSistema).filter(ConfiguracaoSistema.chave == "logo_mime").first()
            if not c_mime:
                db.add(ConfiguracaoSistema(chave="logo_mime", valor=mime_type))
            else:
                c_mime.valor = mime_type
            db.commit()
        except Exception as ex:
            db.rollback()
            print(f"Warning: could not save logo to SQLite DB: {ex}")

        # 3. Save to Firestore for permanent cloud backup across environments
        try:
            from app.firebase_config import get_firestore_client
            db_fs = get_firestore_client()
            if db_fs:
                db_fs.collection("configuracoes").document("logo_data").set({
                    "logo_b64": b64_str,
                    "mime_type": mime_type,
                    "filename": filename,
                    "updated_at": datetime.datetime.utcnow().isoformat()
                })
        except Exception as ex:
            print(f"Warning: could not save logo to Firestore: {ex}")

        # 4. Save to embedded_logo.py for code-level fallback persistence
        try:
            import os
            emb_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "embedded_logo.py")
            with open(emb_file, "w") as ef:
                ef.write('# Auto-generated embedded fallback logo base64 string\n')
                ef.write(f'EMBEDDED_LOGO_B64 = "{b64_str}"\n')
                ef.write(f'EMBEDDED_LOGO_MIME = "{mime_type}"\n')
        except Exception as ex_emb:
            print(f"Notice updating embedded_logo.py: {ex_emb}")

        request.session["msg_success"] = "Logotipo do terreiro atualizado com sucesso!"
    except Exception as e:
        print(f"Error in admin_config_logo: {e}")
        request.session["msg_error"] = f"Erro ao atualizar o logotipo: {e}"

    return RedirectResponse(url="/admin?tab=configuracoes", status_code=303)


@router.post("/admin/config/porcentagem")
async def admin_config_porcentagem(
    request: Request,
    porcentagem_casa: float = Form(...),
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    try:
        from app.models import User, ConfiguracaoSistema
        
        current_user = db.query(User).filter(User.id == user_id).first()
        if not current_user or current_user.role not in ["programador", "pai_de_santo"]:
            request.session["msg_error"] = "Você não tem permissão para alterar as configurações do terreiro."
            return RedirectResponse(url="/admin?tab=configuracoes", status_code=303)
            
        # Try to find existing config, or create it
        config = db.query(ConfiguracaoSistema).filter(ConfiguracaoSistema.chave == "porcentagem_casa").first()
        if not config:
            config = ConfiguracaoSistema(chave="porcentagem_casa", valor=str(porcentagem_casa))
            db.add(config)
        else:
            config.valor = str(porcentagem_casa)
            
        db.commit()
        request.session["msg_success"] = f"Porcentagem destinada à casa atualizada para {porcentagem_casa}% com sucesso!"
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao atualizar porcentagem: {e}"
        
    return RedirectResponse(url="/admin?tab=configuracoes", status_code=303)


@router.post("/admin/usuario/add")
async def admin_user_add(
    request: Request,
    username: str = Form(...),
    email: str = Form(...),
    role: str = Form(...),
    password: str = Form(...),
    confirm_password: str = Form(...),
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    try:
        from app.models import User
        from app.auth_utils import hash_password
        
        current_user = db.query(User).filter(User.id == user_id).first()
        if not current_user or current_user.role not in ["programador", "pai_de_santo"]:
            request.session["msg_error"] = "Acesso negado para cadastrar administradores."
            return RedirectResponse(url="/admin?tab=usuarios", status_code=303)
            
        # Check collision
        existing_user = db.query(User).filter((User.username == username) | (User.email == email)).first()
        if existing_user:
            request.session["msg_error"] = "Já existe um usuário cadastrado com este nome de usuário ou e-mail."
            return RedirectResponse(url="/admin?tab=usuarios", status_code=303)
            
        if password != confirm_password:
            request.session["msg_error"] = "As senhas digitadas não coincidem."
            return RedirectResponse(url="/admin?tab=usuarios", status_code=303)
            
        new_user = User(
            username=username.strip(),
            email=email.strip(),
            hashed_password=hash_password(password),
            is_active=True,
            is_admin=True,
            role=role
        )
        db.add(new_user)
        db.commit()
        request.session["msg_success"] = f"Usuário '{username}' cadastrado com sucesso!"
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao cadastrar usuário: {e}"
        
    return RedirectResponse(url="/admin?tab=usuarios", status_code=303)


@router.post("/admin/usuario/edit/{target_user_id}")
async def admin_user_edit(
    target_user_id: int,
    request: Request,
    username: str = Form(...),
    email: str = Form(...),
    role: str = Form(...),
    password: str = Form(None),
    confirm_password: str = Form(None),
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    try:
        from app.models import User
        from app.auth_utils import hash_password
        
        current_user = db.query(User).filter(User.id == user_id).first()
        if not current_user or current_user.role not in ["programador", "pai_de_santo"]:
            request.session["msg_error"] = "Acesso negado para editar administradores."
            return RedirectResponse(url="/admin?tab=usuarios", status_code=303)
            
        target_user = db.query(User).filter(User.id == target_user_id).first()
        if not target_user:
            request.session["msg_error"] = "Usuário não encontrado."
            return RedirectResponse(url="/admin?tab=usuarios", status_code=303)
            
        # Programadores can only be edited by other Programadores
        if target_user.role == "programador" and current_user.role != "programador":
            request.session["msg_error"] = "Apenas administradores com cargo 'Programador' podem editar contas master."
            return RedirectResponse(url="/admin?tab=usuarios", status_code=303)
            
        # Check collision
        existing_user = db.query(User).filter((User.username == username) | (User.email == email)).filter(User.id != target_user_id).first()
        if existing_user:
            request.session["msg_error"] = "Já existe outro usuário cadastrado com este nome de usuário ou e-mail."
            return RedirectResponse(url="/admin?tab=usuarios", status_code=303)
            
        target_user.username = username.strip()
        target_user.email = email.strip()
        target_user.role = role
        
        if password:
            if password != confirm_password:
                request.session["msg_error"] = "As senhas digitadas não coincidem."
                return RedirectResponse(url="/admin?tab=usuarios", status_code=303)
            target_user.hashed_password = hash_password(password)
            
        db.commit()
        request.session["msg_success"] = f"Usuário '{username}' atualizado com sucesso!"
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao atualizar usuário: {e}"
        
    return RedirectResponse(url="/admin?tab=usuarios", status_code=303)


@router.post("/admin/usuario/delete/{target_user_id}")
async def admin_user_delete(
    target_user_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    try:
        from app.models import User
        
        current_user = db.query(User).filter(User.id == user_id).first()
        if not current_user or current_user.role not in ["programador", "pai_de_santo"]:
            request.session["msg_error"] = "Acesso negado."
            return RedirectResponse(url="/admin?tab=usuarios", status_code=303)
            
        if user_id == target_user_id:
            request.session["msg_error"] = "Você não pode excluir sua própria conta de administrador conectada."
            return RedirectResponse(url="/admin?tab=usuarios", status_code=303)
            
        target_user = db.query(User).filter(User.id == target_user_id).first()
        if not target_user:
            request.session["msg_error"] = "Usuário não encontrado."
            return RedirectResponse(url="/admin?tab=usuarios", status_code=303)
            
        # Programadores can only be deleted by other Programadores
        if target_user.role == "programador" and current_user.role != "programador":
            request.session["msg_error"] = "Apenas administradores com cargo 'Programador' podem excluir contas master."
            return RedirectResponse(url="/admin?tab=usuarios", status_code=303)
            
        username = target_user.username
        db.delete(target_user)
        db.commit()
        request.session["msg_success"] = f"Usuário '{username}' excluído com sucesso."
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao excluir usuário: {e}"
        
    return RedirectResponse(url="/admin?tab=usuarios", status_code=303)


# --------------------------------------------------------------------------
# E. INTERNAL BULLETIN BOARD CONTROLLERS
# --------------------------------------------------------------------------

@router.post("/admin/quadro_avisos/add")
async def quadro_avisos_add(
    request: Request,
    title: str = Form(...),
    content: str = Form(...),
    pub_site: str = Form(None),
    pub_internal: str = Form(None),
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    from app.models import User, Aviso, QuadroAviso
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    permitted_roles = ["programador", "pai_de_santo", "tesoureiro", "secretario"]
    if current_user.role not in permitted_roles:
        request.session["msg_error"] = "Acesso negado para publicar no quadro de avisos."
        return RedirectResponse(url="/admin?tab=quadro_avisos", status_code=303)
        
    try:
        import datetime
        is_site = pub_site == "true"
        is_internal = pub_internal == "true"
        
        # Se nenhum for selecionado, assume interno por padrão
        if not is_site and not is_internal:
            is_internal = True
            
        saved_targets = []
        if is_site:
            # Apenas cargos autorizados podem publicar no site público
            if current_user.role in ["programador", "pai_de_santo", "secretario"]:
                new_aviso = Aviso(
                    title=title,
                    content=content,
                    date_posted=datetime.date.today(),
                    is_active=True
                )
                db.add(new_aviso)
                saved_targets.append("Site Público")
            else:
                # Opcional: ignorar silenciosamente ou avisar que não tem permissão para o site
                pass
                
        if is_internal:
            new_internal = QuadroAviso(
                title=title,
                content=content,
                author_name=current_user.username,
                date_posted=datetime.date.today()
            )
            db.add(new_internal)
            saved_targets.append("Quadro de Avisos Interno")
            
        db.commit()
        targets_str = " e ".join(saved_targets)
        request.session["msg_success"] = f"Aviso publicado com sucesso no {targets_str}!"
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao publicar aviso interno: {e}"
        
    return RedirectResponse(url="/admin?tab=quadro_avisos", status_code=303)


@router.post("/admin/quadro_avisos/delete/{aviso_id}")
async def quadro_avisos_delete(
    aviso_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    from app.models import User
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    permitted_roles = ["programador", "pai_de_santo", "tesoureiro", "secretario"]
    if current_user.role not in permitted_roles:
        request.session["msg_error"] = "Acesso negado para excluir avisos internos."
        return RedirectResponse(url="/admin?tab=quadro_avisos", status_code=303)
        
    try:
        aviso = db.query(QuadroAviso).filter(QuadroAviso.id == aviso_id).first()
        if aviso:
            db.delete(aviso)
            db.commit()
            request.session["msg_success"] = "Aviso interno removido com sucesso!"
        else:
            request.session["msg_error"] = "Aviso interno não encontrado."
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao excluir aviso interno: {e}"
        
    return RedirectResponse(url="/admin?tab=quadro_avisos", status_code=303)


# --------------------------------------------------------------------------
# I. PRIVATE CONSULTATION ACTION CONTROLLERS
# --------------------------------------------------------------------------

@router.post("/admin/consultas/confirm/{booking_id}")
async def admin_consultas_confirm(
    booking_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    from app.models import User, Membro, ConsultaPrivada
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    try:
        booking = db.query(ConsultaPrivada).filter(ConsultaPrivada.id == booking_id).first()
        if not booking:
            request.session["msg_error"] = "Consulta não encontrada."
            return RedirectResponse(url="/admin?tab=consultas", status_code=303)
            
        is_authorized = False
        if current_user.role in ["programador", "pai_de_santo", "secretario"]:
            is_authorized = True
        elif current_user.role == "membro":
            m_rec = db.query(Membro).filter(Membro.email == current_user.email).first()
            if m_rec and booking.membro_id == m_rec.id:
                is_authorized = True
                
        if not is_authorized:
            request.session["msg_error"] = "Acesso negado para gerenciar esta consulta."
            return RedirectResponse(url="/admin?tab=consultas", status_code=303)
            
        booking.status = "Confirmada"
        
        # Extract and save custom consultation fee
        form = await request.form()
        valor_str = form.get("valor_consulta")
        booking.valor_consulta = float(valor_str) if valor_str else 0.0
        
        db.commit()
        request.session["msg_success"] = f"Consulta de {booking.nome_cliente} aceita/confirmada com sucesso com valor de R$ {booking.valor_consulta:.2f}!"
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao confirmar consulta: {e}"
        
    return RedirectResponse(url="/admin?tab=consultas", status_code=303)


@router.post("/admin/consultas/complete/{booking_id}")
async def admin_consultas_complete(
    booking_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    from app.models import User, Membro, ConsultaPrivada, ConfiguracaoSistema, TransacaoFinanceira
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    try:
        booking = db.query(ConsultaPrivada).filter(ConsultaPrivada.id == booking_id).first()
        if not booking:
            request.session["msg_error"] = "Consulta não encontrada."
            return RedirectResponse(url="/admin?tab=consultas", status_code=303)
            
        is_authorized = False
        if current_user.role in ["programador", "pai_de_santo", "secretario"]:
            is_authorized = True
        elif current_user.role == "membro":
            m_rec = db.query(Membro).filter(Membro.email == current_user.email).first()
            if m_rec and booking.membro_id == m_rec.id:
                is_authorized = True
                
        if not is_authorized:
            request.session["msg_error"] = "Acesso negado para gerenciar esta consulta."
            return RedirectResponse(url="/admin?tab=consultas", status_code=303)
            
        booking.status = "Realizada"
        
        # Calculate house revenue cut based on this specific booking's custom value
        membro = db.query(Membro).filter(Membro.id == booking.membro_id).first()
        valor_base = booking.valor_consulta if (booking.valor_consulta is not None and booking.valor_consulta > 0) else (membro.valor_consulta if (membro and membro.valor_consulta > 0) else 0.0)
        
        if valor_base > 0:
            config_pct = db.query(ConfiguracaoSistema).filter(ConfiguracaoSistema.chave == "porcentagem_casa").first()
            porcentagem = float(config_pct.valor) if config_pct else 30.0
            valor_casa = (valor_base * porcentagem) / 100.0
            if valor_casa > 0:
                membro_nome = membro.nome if membro else "Médium"
                # Add financial transaction representing the house revenue portion
                nova_transacao = TransacaoFinanceira(
                    descricao=f"Contribuição de Consulta Privada: {booking.nome_cliente} (Médium: {membro_nome})",
                    valor=valor_casa,
                    tipo="receita",
                    categoria="Consulta Privada",
                    data=datetime.date.today()
                )
                db.add(nova_transacao)
                
        db.commit()
        request.session["msg_success"] = f"Consulta de {booking.nome_cliente} concluída com sucesso!"
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao concluir consulta: {e}"
        
    return RedirectResponse(url="/admin?tab=consultas", status_code=303)


@router.post("/admin/consultas/cancel/{booking_id}")
async def admin_consultas_cancel(
    booking_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    from app.models import User, Membro, ConsultaPrivada
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    try:
        booking = db.query(ConsultaPrivada).filter(ConsultaPrivada.id == booking_id).first()
        if not booking:
            request.session["msg_error"] = "Consulta não encontrada."
            return RedirectResponse(url="/admin?tab=consultas", status_code=303)
            
        is_authorized = False
        if current_user.role in ["programador", "pai_de_santo", "secretario"]:
            is_authorized = True
        elif current_user.role == "membro":
            m_rec = db.query(Membro).filter(Membro.email == current_user.email).first()
            if m_rec and booking.membro_id == m_rec.id:
                is_authorized = True
                
        if not is_authorized:
            request.session["msg_error"] = "Acesso negado para gerenciar esta consulta."
            return RedirectResponse(url="/admin?tab=consultas", status_code=303)
            
        booking.status = "Cancelada"
        db.commit()
        request.session["msg_success"] = f"Consulta de {booking.nome_cliente} recusada/cancelada com sucesso."
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao cancelar consulta: {e}"
        
    return RedirectResponse(url="/admin?tab=consultas", status_code=303)


@router.post("/admin/consultas/delete/{booking_id}")
async def admin_consultas_delete(
    booking_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/admin/login", status_code=303)
        
    from app.models import User
    current_user = db.query(User).filter(User.id == user_id).first()
    if not current_user or current_user.role not in ["programador", "pai_de_santo", "secretario"]:
        request.session["msg_error"] = "Acesso negado para gerenciar consultas."
        return RedirectResponse(url="/admin?tab=consultas", status_code=303)
        
    try:
        booking = db.query(ConsultaPrivada).filter(ConsultaPrivada.id == booking_id).first()
        if booking:
            name = booking.nome_cliente
            db.delete(booking)
            db.commit()
            request.session["msg_success"] = f"Agendamento de {name} excluído do sistema."
        else:
            request.session["msg_error"] = "Consulta não encontrada."
    except Exception as e:
        db.rollback()
        request.session["msg_error"] = f"Erro ao excluir consulta: {e}"
        
    return RedirectResponse(url="/admin?tab=consultas", status_code=303)



